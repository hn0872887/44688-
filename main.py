import pathlib
from openpyxl import Workbook


if __name__ == "__main__":
    PATH = r"D:\Workspaces\DataScience\Kaggle\Respiratory_Sound_Database\audio_and_txt_files"
    RAW_LIST = []
    ROW_LIST = []

    ############
    # GET DATA #
    ############
    data_path = pathlib.Path(PATH)
    txt_files = list(data_path.glob("*.txt"))

    # get data from txt files
    for f in txt_files:
        # file_data = []
        # file_data.append(pathlib.Path(f).stem)
        # for i in pathlib.Path(f).stem.split("_"):
        #     file_data.append(i)

        file_data = pathlib.Path(f).stem.split("_")

        with open(f, 'r') as lines:
            no_crackle = 0
            no_wheeze = 0
            for line in lines:
                no_crackle = no_crackle + int(line.split()[2])
                no_wheeze = no_wheeze + int(line.split()[3])

        file_data.append(no_crackle)
        file_data.append(no_wheeze)

        RAW_LIST.append(file_data)

    ###############
    # REFINE DATA #
    ###############
    key_set = set()
    for item in RAW_LIST:
        key_set.add(item[0])

    for item in sorted(key_set):
        ROW_DATA = []
        patient_id = item
        rec_idx = set()
        chest_loc = set()
        acq = set()
        device = set()
        tot_crackle = 0
        tot_wheeze = 0

        for i in RAW_LIST:
            if i[0] == item:
                rec_idx.add(i[1])
                chest_loc.add(i[2])
                acq.add(i[3])
                device.add(i[4])
                tot_crackle = tot_crackle + int(i[5])
                tot_wheeze = tot_crackle + int(i[6])

        str_rec_idx = ""
        str_chest_loc = ""
        str_acq = ""
        str_device = ""

        for i in rec_idx:
            str_rec_idx = str_rec_idx + i + ", "
            
        for i in chest_loc:
            str_chest_loc = str_chest_loc + i + ", "

        for i in acq:
            str_acq = str_acq + i + ", "

        for i in device:
            str_device = str_device + i + ", "

        ROW_DATA.append(patient_id)
        ROW_DATA.append(str_rec_idx)
        ROW_DATA.append(str_chest_loc)
        ROW_DATA.append(str_acq)
        ROW_DATA.append(str_device)
        ROW_DATA.append(str(tot_crackle))
        ROW_DATA.append(str(tot_wheeze))

        ROW_LIST.append(ROW_DATA)

    #####################
    # PUT DATA TO EXCEL #
    #####################
    current_dir = pathlib.Path.cwd()
    output_file = current_dir / "output_data.xlsx"
    
    if output_file.exists():
        output_file.unlink()    # delete output file if existed

    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1).value = 'PATIENT_ID'
    ws.cell(row=1, column=2).value = 'REC_IDX'
    ws.cell(row=1, column=3).value = 'CHEST_LOC'
    ws.cell(row=1, column=4).value = 'ACQ'
    ws.cell(row=1, column=5).value = 'DEVICE'
    ws.cell(row=1, column=6).value = 'TOT_CRACKLE'
    ws.cell(row=1, column=7).value = 'TOT_WHEEZE'

    count = 1
    for r in ROW_LIST:
        count = count + 1
        ws.cell(row=count, column=1).value = r[0]
        ws.cell(row=count, column=2).value = r[1]
        ws.cell(row=count, column=3).value = r[2]
        ws.cell(row=count, column=4).value = r[3]
        ws.cell(row=count, column=5).value = r[4]
        ws.cell(row=count, column=6).value = r[5]
        ws.cell(row=count, column=7).value = r[6]

    wb.save(output_file)